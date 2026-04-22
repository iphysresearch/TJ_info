#!/usr/bin/env python3
"""
Import Taiji Alliance Institution Data
导入太极联盟成员单位数据

Reads institution data from an Excel file (太极联盟信息整理.xlsx),
merges the two relevant sheets, applies data corrections, and outputs
a structured JSON file.

Usage:
    python scripts/import_institutions.py --input /path/to/太极联盟信息整理.xlsx
    python scripts/import_institutions.py --input /path/to/file.xlsx --output data/institutions.json
"""

import sys
import json
import argparse
import re
import logging
from pathlib import Path
from datetime import date
from typing import Optional

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# ─────────────────────────────────────────────────────────────────────────────
# English name mapping for the 37 institutions missing English names
# ─────────────────────────────────────────────────────────────────────────────
ENGLISH_NAME_MAP = {
    "中国科学院南京天文光学技术研究所": "Nanjing Institute of Astronomical Optics & Technology, CAS",
    "Mississippi State /West Lake University": "Mississippi State / Westlake University",
    "大连理工大学": "Dalian University of Technology",
    "复旦大学": "Fudan University",
    "海南热带海洋学院": "Hainan Tropical Ocean University",
    "杭州电子科技大学": "Hangzhou Dianzi University",
    "杭州师范大学": "Hangzhou Normal University",
    "湖南师范大学": "Hunan Normal University",
    "华中师范大学": "Central China Normal University",
    "吉林大学": "Jilin University",
    "空间科学与应用研究中心": "Center for Space Science and Applied Research",
    "李政道研究所／台湾大学": "T.D. Lee Institute / National Taiwan University",
    "辽宁师范大学": "Liaoning Normal University",
    "南华大学": "University of South China",
    "南开大学": "Nankai University",
    "南洋理工大学": "Nanyang Technological University",
    "清华大学": "Tsinghua University",
    "山西大学": "Shanxi University",
    "上海交通大学": "Shanghai Jiao Tong University",
    "台湾清华大学": "National Tsing Hua University",
    "同济大学": "Tongji University",
    "武汉大学": "Wuhan University",
    "烟台大学": "Yantai University",
    "扬州大学": "Yangzhou University",
    "郑州大学": "Zhengzhou University",
    "中电26所": "CETC 26th Research Institute",
    "中国计量科学研究院": "National Institute of Metrology, China",
    "中国科学技术大学": "University of Science and Technology of China",
    "中国科学院工程热物理研究所": "Institute of Engineering Thermophysics, CAS",
    "中国科学院国家授时中心": "National Time Service Center, CAS",
    "中国科学院合肥物质科学研究院": "Hefei Institutes of Physical Science, CAS",
    "中国科学院软件研究所": "Institute of Software, CAS",
    "中国科学院上海微系统与信息技术研究所": "Shanghai Institute of Microsystem and Information Technology, CAS",
    "中国科学院苏州医工所": "Suzhou Institute of Biomedical Engineering and Technology, CAS",
    "中国科学院武汉物理与数学研究所": "Wuhan Institute of Physics and Mathematics, CAS",
    "中国科学院云南天文台": "Yunnan Observatories, CAS",
    "重庆大学": "Chongqing University",
    "重庆声光电有限公司": "Chongqing Acousto-Optoelectronics Co., Ltd.",
}

# Typo corrections for the 成员单位英文 sheet matching
# "南阳理工大学" in that sheet actually refers to "南洋理工大学"
PARTNER_NAME_CORRECTIONS = {
    "南阳理工大学": "南洋理工大学",
}

# English name typo corrections
ENGLISH_NAME_CORRECTIONS = {
    "Nanyang TechnologicalUniversity": "Nanyang Technological University",
}


def make_slug(name_en: str) -> str:
    """Generate a URL-friendly slug from an English name."""
    slug = name_en.lower()
    # Remove common suffixes/noise
    slug = slug.replace(", cas", "").replace(",cas", "")
    slug = slug.replace(", cast", "").replace(",cast", "")
    slug = slug.replace(", ucas", "").replace(",ucas", "")
    slug = slug.replace(", china", "")
    slug = slug.replace("co., ltd.", "co-ltd")
    slug = slug.replace("&", "and")
    # Replace non-alphanumeric with hyphens
    slug = re.sub(r'[^a-z0-9]+', '-', slug)
    # Clean up multiple/leading/trailing hyphens
    slug = re.sub(r'-+', '-', slug).strip('-')
    return slug


def parse_cooperation_types(text: Optional[str]) -> list:
    """Parse cooperation types string into a list."""
    if not text or not isinstance(text, str):
        return []
    # Split on Chinese comma or 、
    parts = re.split(r'[、，,]', text)
    return [p.strip() for p in parts if p.strip()]


def load_excel(input_path: str) -> tuple:
    """Load both sheets from the Excel file using openpyxl."""
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl is required. Install with: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(input_path, data_only=True)

    # Sheet 1: 成员单位 (65 rows, 2 cols: name_zh, name_en)
    ws_members = wb['成员单位']
    members = []
    for row in ws_members.iter_rows(values_only=True):
        name_zh = str(row[0]).strip() if row[0] else None
        name_en = str(row[1]).strip() if row[1] else None
        if name_zh:
            members.append({
                'name_zh': name_zh,
                'name_en': name_en if name_en and name_en != 'None' else None,
            })

    # Sheet 2: 成员单位英文 (skip header row, 29 data rows)
    ws_partners = wb['成员单位英文']
    partners = {}
    first_row = True
    for row in ws_partners.iter_rows(values_only=True):
        if first_row:
            first_row = False
            continue
        name_zh = str(row[1]).strip() if row[1] else None
        name_en = str(row[2]).strip() if row[2] else None
        cooperation = str(row[3]).strip() if row[3] else None
        if name_zh:
            # Apply name corrections for matching
            corrected_name = PARTNER_NAME_CORRECTIONS.get(name_zh, name_zh)
            # Apply English name corrections
            if name_en:
                name_en = ENGLISH_NAME_CORRECTIONS.get(name_en, name_en)
            partners[corrected_name] = {
                'name_en': name_en,
                'cooperation': cooperation,
            }

    logger.info(f"Loaded {len(members)} members and {len(partners)} partners from Excel")
    return members, partners


def merge_data(members: list, partners: dict) -> list:
    """Merge member list with partner details."""
    entries = []

    for member in members:
        name_zh = member['name_zh']
        name_en = member['name_en']

        # Check if this institution is a partner
        is_partner = name_zh in partners
        partner_info = partners.get(name_zh, {})

        # Determine English name (priority: member sheet > partner sheet > mapping)
        if name_en:
            final_en = name_en
        elif partner_info.get('name_en'):
            final_en = partner_info['name_en']
        elif name_zh in ENGLISH_NAME_MAP:
            final_en = ENGLISH_NAME_MAP[name_zh]
        else:
            logger.warning(f"No English name found for: {name_zh}")
            final_en = name_zh  # Fallback to Chinese name

        # Apply English name corrections on final result too
        final_en = ENGLISH_NAME_CORRECTIONS.get(final_en, final_en)

        entry_id = make_slug(final_en)

        entry = {
            'entry_id': entry_id,
            'name_zh': name_zh,
            'name_en': final_en,
            'institution_type': 'partner' if is_partner else 'affiliate',
        }

        if is_partner:
            entry['cooperation_types'] = parse_cooperation_types(
                partner_info.get('cooperation')
            )
        else:
            entry['cooperation_types'] = None

        entries.append(entry)

    return entries


def check_duplicates(entries: list) -> list:
    """Check for duplicate entry_ids and resolve by appending suffix."""
    seen = {}
    for entry in entries:
        eid = entry['entry_id']
        if eid in seen:
            # Append a numeric suffix
            i = 2
            while f"{eid}-{i}" in seen:
                i += 1
            new_id = f"{eid}-{i}"
            logger.warning(
                f"Duplicate entry_id '{eid}' for '{entry['name_zh']}', "
                f"renamed to '{new_id}'"
            )
            entry['entry_id'] = new_id
            seen[new_id] = entry
        else:
            seen[eid] = entry
    return entries


def build_output(entries: list, data_source: str) -> dict:
    """Build the final JSON structure."""
    today = date.today().isoformat()
    n_partners = sum(1 for e in entries if e['institution_type'] == 'partner')
    n_affiliates = sum(1 for e in entries if e['institution_type'] == 'affiliate')

    return {
        "metadata": {
            "version": "1.0",
            "project": "Taiji Alliance Institutions",
            "created_date": today,
            "last_updated": today,
            "total_entries": len(entries),
            "description": "太极联盟成员单位数据库",
            "data_source": data_source,
            "statistics": {
                "official_partners": n_partners,
                "affiliate_institutions": n_affiliates,
            }
        },
        "entries": entries,
    }


def main():
    parser = argparse.ArgumentParser(
        description="Import Taiji Alliance institution data from Excel"
    )
    parser.add_argument(
        '--input', '-i',
        required=True,
        help='Path to the Excel file (太极联盟信息整理.xlsx)'
    )
    parser.add_argument(
        '--output', '-o',
        default='data/institutions.json',
        help='Output JSON path (default: data/institutions.json)'
    )
    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        logger.error(f"Input file not found: {input_path}")
        sys.exit(1)

    output_path = Path(args.output)

    # Load data
    members, partners = load_excel(str(input_path))

    # Merge
    entries = merge_data(members, partners)

    # Check for duplicates
    entries = check_duplicates(entries)

    # Build output
    data = build_output(entries, input_path.name)

    # Write
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    logger.info(f"Wrote {len(entries)} entries to {output_path}")

    # Summary
    n_partners = data['metadata']['statistics']['official_partners']
    n_affiliates = data['metadata']['statistics']['affiliate_institutions']
    logger.info(f"  Partners: {n_partners}, Affiliates: {n_affiliates}")

    # Check all have English names
    missing = [e for e in entries if e['name_en'] == e['name_zh']]
    if missing:
        logger.warning(f"  {len(missing)} entries still missing English names:")
        for e in missing:
            logger.warning(f"    - {e['name_zh']}")


if __name__ == '__main__':
    main()
